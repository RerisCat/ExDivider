# This file is part of ExDivider.
#
# ExDivider is free software: you can redistribute it and/or modify 
# it under the terms of the GNU General Public License as published 
# by the Free Software Foundation, either version 3 of the License, 
# or (at your option) any later version.
#
# ExDivider is distributed in the hope that it will be useful, but 
# WITHOUT ANY WARRANTY; without even the implied warranty of 
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU 
# General Public License for more details.
#
# You should have received a copy of the GNU General Public License 
# along with Foobar. If not, see <https://www.gnu.org/licenses/>.

import typing

from openpyxl.utils import get_column_letter

from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import QWidget, QLineEdit, QAbstractItemView, QTabBar
from PyQt5.QtCore import QModelIndex, Qt, QAbstractTableModel, QVariant


class SheetModel(QAbstractTableModel):
    def __init__(self, sheet=[[]], parent=None):
        super(SheetModel, self).__init__(parent)
        self.sheet = sheet
        self.indeces_conn_excel = []
        self.indeces_default_value = []

    def headerData(self, section: int, orientation: Qt.Orientation, role: int):
        if role == QtCore.Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return get_column_letter(section + 1)
            else:
                return str(section + 1)

    def columnCount(self, parent=None):
        return self.sheet.max_column

    def rowCount(self, parent=None):
        return self.sheet.max_row

    def setData(self, index: QModelIndex, value: QVariant, role: int):
        if role == Qt.DisplayRole:
            row = index.row() + 1
            col = index.column() + 1
            self.sheet[get_column_letter(col)+str(row)] = value.value()
        
        return super(SheetModel, self).setData(index, value, role)
    
    def setSheet(self, sheet):
        self.sheet = sheet
    
    def addIndecesConnExcel(self, index):
        if index not in self.indeces_conn_excel:
            self.indeces_conn_excel.append(index)
    def removeIndecesConnExcel(self, index):
        if index in self.indeces_conn_excel:
            self.indeces_conn_excel.remove(index)
        
    def addIndecesDefaultValue(self, index):
        if index not in self.indeces_default_value:
            self.indeces_default_value.append(index)
    def removeIndecesDefaultValue(self, index):
        if index in self.indeces_default_value:
            self.indeces_default_value.remove(index)

    def data(self, index: QModelIndex, role: int):
        if role == Qt.DisplayRole:
            row = index.row() + 1
            col = index.column() + 1
            value = self.sheet[get_column_letter(col)+str(row)].value
            if value is not None:
                return str(value)
            else:
                return ''
        if role == QtCore.Qt.BackgroundColorRole:
            if index in self.indeces_conn_excel and index in self.indeces_default_value:
                gradient = QtGui.QLinearGradient(0, 0, 60, 50)
                gradient.setColorAt(0, Qt.green)
                gradient.setColorAt(1, QtGui.QColor( 0xFF, 0xA0, 0x00 ))
                
                brush = QtGui.QBrush(gradient)
                return brush
            elif index in self.indeces_conn_excel:
                brush = QtGui.QBrush(Qt.green)
                return brush
            elif index in self.indeces_default_value:
                brush = QtGui.QBrush(QtGui.QColor( 0xFF, 0xA0, 0x00 ))
                return brush
            
class DialogDragItemsMenu(QWidget):
    items_ready = QtCore.pyqtSignal(dict)

    def __init__(self, parent=None):
        super(DialogDragItemsMenu, self).__init__(parent)
        
        self.setWindowFlags(
            self.windowFlags() | QtCore.Qt.Popup | QtCore.Qt.FramelessWindowHint)
        self.setMaximumHeight(400)
        self.setMinimumWidth(500)
        self.setMaximumWidth(500)
        
        layout = QtWidgets.QVBoxLayout(self)
        
        label_dialog = QtWidgets.QLabel(self)
        label_dialog.setText('Порядок ввода элементов:')
        layout.addWidget(label_dialog)

        self.dialogList = QtWidgets.QListWidget()
        self.dialogList.horizontalScrollBar().setEnabled(False);
        self.dialogList.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.dialogList.setFocusPolicy(Qt.NoFocus)
        
        self.dialogList.setDragDropMode(QAbstractItemView.InternalMove)
        self.dialogList.clicked.connect(self.test)

        layout.addWidget(self.dialogList)
        
        hlayout_le_wrapper = QWidget()
        hlayout_le = QtWidgets.QHBoxLayout(hlayout_le_wrapper)
        label_del = QtWidgets.QLabel(self)
        label_del.setText('Введите разделитель:')
        hlayout_le.addWidget(label_del)
        self.le_delimiter = QLineEdit(self)
        hlayout_le.addWidget(self.le_delimiter)
        layout.addWidget(hlayout_le_wrapper)
        
        label_warning = QtWidgets.QLabel(self)
        label_warning.setText('*Разделитель – символ или символы, разделяющие ячейки. По умолчанию новая ячейка будет введена без пробела')
        label_warning.setWordWrap(True)
        layout.addWidget(label_warning)
        
        hlayout_buttons_wrapper = QWidget()
        hlayout_buttons = QtWidgets.QHBoxLayout(hlayout_buttons_wrapper)
        button_ok = QtWidgets.QPushButton(self)
        button_ok.setText('Выбрать')
        button_ok.clicked.connect(self.click_select)
        hlayout_buttons.addWidget(button_ok)
        button_close = QtWidgets.QPushButton(self)
        button_close.setText('Закрыть')
        button_close.clicked.connect(self.hide)
        hlayout_buttons.addWidget(button_close)
        layout.addWidget(hlayout_buttons_wrapper)
    
    def test(self):
        print([self.dialogList.item(i).text() for i in range(self.dialogList.count())])
                
    def add_items(self, items):
        self.dialogList.clear()
        self.dialogList.addItems(items)
        
    def show_dialog(self, pos):
        self.posit = pos
        self.move(pos)
        self.show()
        
    def click_select(self):
        items_items = [self.dialogList.item(i).text() for i in range(self.dialogList.count())]
        
        items_letter = []
        for item in items_items:
            (firstWord, rest) = str(item).split(maxsplit=1)
            items_letter.append(firstWord)
        items = {
            'items' : items_letter,
            'delimeter' : str(self.le_delimiter.text())
        }
        self.items_ready.emit(items)
        self.close()
            
class DialogWidgetMultipleMenu(QWidget):
    items_ready = QtCore.pyqtSignal(dict)

    def __init__(self, parent=None):
        super(DialogWidgetMultipleMenu, self).__init__(parent)
        
        self.setWindowFlags(
            self.windowFlags() | QtCore.Qt.Popup | QtCore.Qt.FramelessWindowHint)
        self.setMaximumHeight(300)
        self.setMinimumWidth(500)
        
        layout = QtWidgets.QVBoxLayout(self)

        self.dialogList = QtWidgets.QListWidget()
        self.dialogList.horizontalScrollBar().setEnabled(False);
        self.dialogList.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.dialogList.setSelectionMode(QAbstractItemView.MultiSelection)
        layout.addWidget(self.dialogList)
        
        hlayout_wrapper = QWidget()

        hlayout = QtWidgets.QHBoxLayout(hlayout_wrapper)
        button_ok = QtWidgets.QPushButton(self)
        button_ok.setText('Выбрать')
        button_ok.clicked.connect(self.click_select)
        hlayout.addWidget(button_ok)
        button_close = QtWidgets.QPushButton(self)
        button_close.setText('Закрыть')
        button_close.clicked.connect(self.hide)
        hlayout.addWidget(button_close)
        
        layout.addWidget(hlayout_wrapper)
        
        self.table_view_dialog_drag = DialogDragItemsMenu()
        self.table_view_dialog_drag.items_ready.connect(self.pass_items)
                
    def add_items(self, items):
        self.dialogList.clear()
        self.dialogList.addItems(items[1:])
        
    def show_dialog(self, pos):
        self.posit = pos
        self.move(pos)
        self.show()
        
    def click_select(self):
        items = [item.text() for item in self.dialogList.selectedItems()]
        if len(items) > 0:
            self.table_view_dialog_drag.add_items(items) 
            self.table_view_dialog_drag.show_dialog(self.posit)
            self.close()
        
    def pass_items(self, dict_in):
        self.items_ready.emit(dict_in)
            
class DialogWidgetMenu(QWidget):
    clicked = QtCore.pyqtSignal(int)

    def __init__(self, parent=None):
        super(DialogWidgetMenu, self).__init__(parent)
        
        self.setWindowFlags(
            self.windowFlags() | QtCore.Qt.Popup | QtCore.Qt.FramelessWindowHint)
        self.setMaximumHeight(300)
        self.setMinimumWidth(500)
        
        layout = QtWidgets.QVBoxLayout(self)

        self.dialogList = QtWidgets.QListWidget()
        self.dialogList.horizontalScrollBar().setEnabled(False);
        self.dialogList.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.dialogList.clicked.connect(self.slot_select)
        layout.addWidget(self.dialogList)
        
        buttonBox = QtWidgets.QPushButton(self)
        buttonBox.setText('Закрыть')
        buttonBox.clicked.connect(self.hide)
        layout.addWidget(buttonBox)
                
    def slot_select(self, item):
        self.clicked.emit(item.row())
        self.close()
        
    def add_items(self, items):
        self.dialogList.clear()
        self.dialogList.addItems(items)
        
    def show_dialog(self, pos):
        self.move(pos)
        self.show()

class WheelBar(QTabBar):
    def __init__(self, parent: typing.Optional[QWidget] = ...) -> None:
        super().__init__()
        self.setShape(self.Shape.RoundedSouth)
    
    def wheelEvent(self, event: QtGui.QWheelEvent) -> None:
        event.accept()
        
class LicenseWindow(QWidget):

    def __init__(self, icon, parent=None):
        self.parent = parent
        QWidget.__init__(self)
        
        self.setWindowModality(QtCore.Qt.ApplicationModal)
        
        self.setWindowIcon(icon)
        self.setWindowTitle("О программе")
        
        self.setMaximumHeight(300)
        self.setMinimumWidth(500)
        
        layout = QtWidgets.QVBoxLayout(self)

        self.short_license = QtWidgets.QLabel()
        self.short_license.setText("ExDivider версии 1.0, Copyright © 2022; Ядрихинский Николай(RerisCat). \n\n" +
                                   "ExDivider распространяется БЕЗО ВСЯКИХ ГАРАНТИЙ. Эта программа распространяется \n"+
                                   "на условиях Стандартной общественной лицензии GNU (GNU GPL v3). \n\n"+
                                   "Для создания были использованы библиотеки: \n"+
                                   "    -openpyxl, лицензия MIT License\n"+
                                   "    -PyInstaller, лицензия GNU GPL v2\n"+
                                   "    -Qt5, лицензия GNU GPL v2\n"+
                                   "    -PyQt5, лицензия GNU GPL v3\n\n"+
                                   "Исходный код: https://github.com/RerisCat/ExDivider")

        layout.addWidget(self.short_license)
        
        button_close = QtWidgets.QPushButton(self)
        button_close.setText('Закрыть')
        button_close.clicked.connect(self.hide)
        layout.addWidget(button_close)
                
    def showEvent(self, event):
        if not event.spontaneous():
            geo = self.geometry()
            geo.moveCenter(self.parent.geometry().center())
            QtCore.QTimer.singleShot(0, lambda: self.setGeometry(geo))
        