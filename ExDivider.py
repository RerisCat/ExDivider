# This file is part of ExDivider.
#
# ExDivider is free software: you can redistribute it and/or modify 
# it under the terms of the GNU General Public License as published 
# by the Free Software Foundation, either version 3 of the License, 
# or (at your option) any later version.
#
# ExDivider is distributed in the hope that it will be useful, but 
# WITHOUT ANY WARRANTY; without even the implied warranty of 
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU 
# General Public License for more details.
#
# You should have received a copy of the GNU General Public License 
# along with Foobar. If not, see <https://www.gnu.org/licenses/>.

from asyncio.windows_events import NULL

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QListView, QWidget, QMenu, QAction,\
    QInputDialog, QShortcut, QMessageBox
from PyQt5.QtCore import Qt, QStringListModel, QVariant, QFileInfo

import sys, os
import datetime
import json
from pathlib import Path
from configparser import ConfigParser
from table import Ui_MainWindow

from Components import DialogWidgetMenu, DialogWidgetMultipleMenu, LicenseWindow, WheelBar, SheetModel

def replace_illegible_chars(value, deletechars):
    for c in deletechars:
        value = value.replace(c, '_')
    return value;

class Window(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(Window, self).__init__()
        self.setupUi(self)
        
        self.icon = QtGui.QIcon(os.path.dirname(os.path.abspath(__file__)) + '/exdiv_icon.ico')
        
        self.setFixedSize(self.size())   
        self.setWindowIcon(self.icon) 
             
        self.reset()
        
        self.button_select_source.clicked.connect(self.load_source_file)
        self.button_select_pattern.clicked.connect(self.load_pattern_file)
        self.button_start.clicked.connect(self.divide_source_file)
        
        self.action_new.triggered.connect(self.action_new_handler)
        self.action_save.triggered.connect(self.action_save_handler)
        self.action_save_as.triggered.connect(self.action_save_as_handler)
        self.action_load.triggered.connect(self.action_load_handler)
        self.action_about.triggered.connect(self.action_about_handler)
        self.shortcut = QShortcut(QtGui.QKeySequence("Ctrl+S"), self)
        self.shortcut.activated.connect(self.action_save_handler)

        self.table_view.clicked.connect(self.bind_cell)
        self.table_view.setFocusPolicy(Qt.NoFocus)
        self.table_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(self.table_context_menu)
        
        self.line_edit_source.returnPressed.connect(self.line_source_edited)
        self.line_edit_pattern.returnPressed.connect(self.line_pattern_edited)

        # select dialog
        self.table_view_dialog = DialogWidgetMenu()
        self.table_view_dialog.clicked.connect(self.set_cell_name)
        
        self.table_view_dialog_multiple = DialogWidgetMultipleMenu()
        self.table_view_dialog_multiple.items_ready.connect(self.set_cell_mult_names)
        #
        self.license = LicenseWindow(self.icon, self)
        
        self.tabWidget.setTabBar(WheelBar())
        
        listView = QListView()
        listView.setWordWrap(True)
        self.combo_box_cell.setView(listView)
        self.combo_box_cell.currentIndexChanged.connect(self.source_col_changed)
        
        if Path("settings.ini").is_file():
            config = ConfigParser()
            config.read('settings.ini')
            
            save_path = config.get('Info', 'recent_save')
            
            if Path("saves/" + str(Path(save_path).name)).exists():
                save_path = "saves/" + str(Path(save_path).name)
                
            if Path(save_path).is_file():
                self.action_load_handler(save_path)
            else:
                self.showWarning("Сохранения " + save_path +" не существует")
                
    def line_source_edited(self):
        if self.source_path == self.line_edit_source.text():
            return

        self.process_source_file(self.line_edit_source.text())
        
    def line_pattern_edited(self):
        if self.pattern_path == self.line_edit_pattern.text():
            return

        self.process_pattern_file(self.line_edit_pattern.text())
        
    def table_context_menu(self, pos):
        index = self.table_view.indexAt(pos)
        posit = self.table_view.viewport().mapToGlobal(pos)
        menu = QMenu(self)
        
        action_add_multiple = QAction('Добавить несколько значений', self)
        action_add_multiple.triggered.connect(lambda: self.input_several_values(index, posit))
        menu.addAction(action_add_multiple)
        
        action_add_def_value = QAction('Добавить значение по умолчанию', self)
        action_add_def_value.triggered.connect(lambda: self.input_default_value(index))
        menu.addAction(action_add_def_value)
        
        action_remove_def_value = QAction('Удалить значение по умолчанию', self)
        action_remove_def_value.triggered.connect(lambda: self.remove_default_value(index))
        menu.addAction(action_remove_def_value)
        
        menu.popup(posit)
        
    def input_several_values(self, index, pos):
        self.active_cell = [index.row(), index.column()]
        self.table_view_dialog_multiple.show_dialog(pos)
        
    def input_default_value(self, index):
        self.active_cell = [index.row(), index.column()]
        text, ok = QInputDialog.getText(self, 'Ввод', 'Введите значение по умолчанию')
        if ok:
            self.add_data_to_template(self.excels_default_value, text, True)
            sheet_model = self.pattern_sheets[self.active_sheet_name]
            sheet_model.addIndecesDefaultValue(index)
            sheet_model.endResetModel()
            
    def remove_default_value(self, index):
        self.active_cell = [index.row(), index.column()]
        self.add_data_to_template(self.excels_default_value, None, False)
        sheet_model = self.pattern_sheets[self.active_sheet_name]
        sheet_model.removeIndecesDefaultValue(index)
        sheet_model.endResetModel()
        
    def display_sheet(self, sheet_name):  
        self.active_sheet_name = sheet_name
        self.table_view.setModel(self.pattern_sheets[sheet_name])
        
        self.table_view.clearSpans()
        for merged_cell in self.pattern_wb[sheet_name].merged_cells.ranges:
            cells = str(merged_cell).split(':')
            beg_cell = coordinate_from_string(cells[0])
            end_cell = coordinate_from_string(cells[1])
            
            beg_row = beg_cell[1] - 1
            beg_column = column_index_from_string(beg_cell[0]) - 1
            end_row_span = end_cell[1] - beg_row
            end_column_span = column_index_from_string(end_cell[0]) - beg_column
            
            self.table_view.setSpan(beg_row, beg_column ,end_row_span, end_column_span)
            
    def source_col_changed(self, value):
        if value != 0:
            self.current_source_column = get_column_letter(value)
        
    def bind_cell(self, cell):
        self.active_cell = [cell.row(), cell.column()]
        self.table_view_dialog.show_dialog(QtGui.QCursor.pos())

    def set_cell_name(self, row):
        if len(self.active_cell) < 1 or self.active_sheet_name == '':
            return
        sheet_model = self.pattern_sheets[self.active_sheet_name]
        cell_index = sheet_model.index(self.active_cell[0], self.active_cell[1])
        cell_value = ''
        if row != 0:
            cell_value = QVariant(self.source_col_names[row])  
            sheet_model.addIndecesConnExcel(cell_index)
            self.add_data_to_template(self.excels_relation_many, None, False)
            self.add_data_to_template(self.excels_relation, get_column_letter(row), True)
        else:
            sheet = self.orig_pattern_wb[self.active_sheet_name]
            act_row = self.active_cell[0] + 1
            act_col = self.active_cell[1] + 1
            cell_value = QVariant(sheet[get_column_letter(act_col)+str(act_row)].value) 
            sheet_model.removeIndecesConnExcel(cell_index)
            self.add_data_to_template(self.excels_relation_many, None, False)
            self.add_data_to_template(self.excels_relation, None, False)

        sheet_model.setData(cell_index, cell_value, Qt.DisplayRole)
        sheet_model.endResetModel()
        
    def set_cell_mult_names(self, dict_in):
        if len(self.active_cell) < 1:
            return
        sheet_model = self.pattern_sheets[self.active_sheet_name]
        cell_index = sheet_model.index(self.active_cell[0], self.active_cell[1])
        items = dict_in['delimeter'].join(x for x in dict_in['items'])
        
        cell_value = QVariant(items)  
        sheet_model.addIndecesConnExcel(cell_index)
        self.add_data_to_template(self.excels_relation, None, False)
        self.add_data_to_template(self.excels_relation_many, dict_in, True)

        sheet_model.setData(cell_index, cell_value, Qt.DisplayRole)
        sheet_model.endResetModel()
        
    def add_data_to_template(self, dictionary_to, value, add):
        if len(self.active_cell) < 1:
            return
        act_row = self.active_cell[0] + 1
        act_col = self.active_cell[1] + 1
        cell_excel_string = get_column_letter(act_col)+str(act_row)
        
        if add:
            dictionary_to[self.active_sheet_name][cell_excel_string] = value
        elif cell_excel_string in dictionary_to[self.active_sheet_name]:
            del dictionary_to[self.active_sheet_name][cell_excel_string]
                    
    def load_source_file(self):
        path, _ = QFileDialog.getOpenFileName(self, 'Откройте excel файл', str(Path().resolve()), '*.xlsx; *.xlsm; *.xltx; *.xltm')
        if path == '' or self.source_path == path:
            return
                
        self.process_source_file(path)
        
    def process_source_file(self, path, showWarning = True):
        path = str(Path(path).resolve())
        
        if(self.is_excel_file(path)):
            self.source_path = path
            self.line_edit_source.setText(path)
            wb = load_workbook(path, data_only=True)
            sheets = wb.sheetnames
            self.source_ws = wb[sheets[0]]
            self.progress_bar.setValue(0)
            
            self.refresh_list_items()
            return True
        else:
            if showWarning:
                self.showWarning("Файл " + path + " не является excel файлом")
            return False
        
    def load_pattern_file(self):
        path, _ = QFileDialog.getOpenFileName(self, 'Откройте excel файл', str(Path().resolve()), '*.xlsx; *.xlsm; *.xltx; *.xltm')
        if path == '' or self.pattern_path == path:
            return
        
        self.process_pattern_file(path)
        
    def process_pattern_file(self, path: str, showWarning = True):
        path = str(Path(path).resolve())
        
        if(self.is_excel_file(path)):
            self.excels_default_value = {}
            self.excels_relation = {}
            self.excels_relation_many = {}
            self.active_cell = []
            self.pattern_path = path
            self.line_edit_pattern.setText(path)
            self.orig_pattern_wb = load_workbook(path)
            self.pattern_wb = load_workbook(path)
            self.tabs_names = self.pattern_wb.sheetnames
            for name in self.tabs_names:
                self.excels_default_value[name] = {}           
                self.excels_relation[name] = {}
                self.excels_relation_many[name] = {}
            
            self.refresh_tabs()
            self.progress_bar.setValue(0)
            self.refresh_table()
            return True
        else:
            if showWarning:
                self.showWarning("Файл " + path + " не является excel файлом")
            return False
        
    def is_excel_file(self, path: str) -> bool:
        if(path != '' and Path(path).exists()):
            return (path.endswith('.xlsx') or path.endswith('.xlsm') or path.endswith('.xltx') or path.endswith('.xltm'))
        else:
            return False
        
    def refresh_tabs(self):
        for _ in range(self.tabWidget.count()):
            self.tabWidget.removeTab(0)
            
        for name in self.tabs_names:
            tab = QWidget()
            tab.setObjectName(name)
            self.tabWidget.addTab(tab, name)
            self.tabWidget.tabBarClicked.connect(self.tab_clicked)
            
            sheet = self.pattern_wb[name] 
            self.pattern_sheets[name] = SheetModel(sheet)
            
        if(self.tabs_names):
            self.display_sheet(self.tabs_names[0])
    
    def tab_clicked(self, index):
        self.display_sheet(self.tabs_names[index])
            
    def refresh_list_items(self):
        self.source_col_names = []
        self.source_col_names.append('Не использовать')
        if(self.source_ws and str(self.source_ws) != ''):
            for col in range(1, self.source_ws.max_column + 1):
                col_letter = get_column_letter(col)
                if col_letter is not None and (self.source_ws[col_letter+'1'].value is not None or self.source_ws[col_letter+'1'].value != '' ):
                    self.source_col_names.append(col_letter + ' ' + str(self.source_ws[col_letter+'1'].value))
                elif col_letter is not None:
                    self.source_col_names.append(col_letter)
        
        self.table_view_dialog.add_items(self.source_col_names)
        self.table_view_dialog_multiple.add_items(self.source_col_names)
        
        self.combo_box_cell.setModel(QStringListModel(self.source_col_names))
        self.combo_box_cell.setCurrentIndex(0)
        self.refresh_table()
        
    def divide_source_file(self):
        self.progress_bar.setValue(0)
        if self.source_path == '' or self.pattern_path == '':
            msgWarning = QMessageBox()
            msgWarning.setText("Для начала работы нужно чтобы поля источник информации и шаблон были заполнены"); 
            msgWarning.setIcon(QMessageBox.Icon.Information);
            msgWarning.setWindowTitle("Информация");
            msgWarning.setWindowIcon(self.icon) 
            msgWarning.exec();
            return
        if self.combo_box_cell.currentIndex() == 0 and not self.check_box_added.isChecked():
            msgWarning = QMessageBox()
            msgWarning.setText('Нужно выбрать поле "Использовать столбец как новые имена"'); 
            msgWarning.setIcon(QMessageBox.Icon.Information);
            msgWarning.setWindowTitle("Информация");
            msgWarning.setWindowIcon(self.icon) 
            msgWarning.exec();
            return
        
        Path('files').mkdir(parents=True, exist_ok=True)
        for i, row in enumerate(self.source_ws.iter_rows(min_row=2, values_only=True)):
            for sheet in self.pattern_wb:
                default_dict = self.excels_default_value[sheet.title]
                for key in default_dict.keys():
                    sheet[key] = default_dict[key]
                    
                sheet_dict = self.excels_relation[sheet.title]
                for key in sheet_dict.keys():
                    value = ''
                    if len(row) < column_index_from_string(sheet_dict[key]):
                        value = None
                    else:
                        value = row[column_index_from_string(sheet_dict[key]) - 1]
                    
                    if value is not None and str(value) != '':
                        if isinstance(value, datetime.datetime):
                            sheet[key] = str(value.strftime('%d.%m.%Y'))
                        else:
                            sheet[key] = str(value)
                    elif key not in default_dict.keys():
                        orig_sheet = self.orig_pattern_wb[sheet.title]
                        sheet[key] = orig_sheet[key].value
                
                sheet_many_dict = self.excels_relation_many[sheet.title]
                for key in sheet_many_dict.keys():
                    comb_dict = sheet_many_dict[key]
                    values = []
                    for item in comb_dict['items']:
                        value = ''
                        if len(row) < column_index_from_string(item) + 1:
                            value = None
                        else:
                            value = row[column_index_from_string(item) - 1]
                        
                        if isinstance(value, datetime.datetime):
                            value = str(value.strftime('%d.%m.%Y'))
                        else:
                            value = str(value)
                        if value != 'None' and value != '':
                            values.append(value)
                    values = comb_dict['delimeter'].join(x for x in values)
                    sheet[key] = values
                        
            filename = ''
            if self.combo_box_cell.currentIndex() != 0:
                firstWord = str(self.combo_box_cell.currentText()).split(' ')[0]
                column_index = column_index_from_string(firstWord) - 1
                filename += replace_illegible_chars(str(row[column_index]), '\/:*?"<>|') 
            
            if self.check_box_added.isChecked() or filename is None or filename.replace(' ', '') == '':
                filename += str(i)
            if str(self.line_edit_added.text()) != '':
                filename += str(self.line_edit_added.text())
            filename = filename.replace('\n', ' ')
            
            try:
                self.pattern_wb.save('files/' + filename + '.xlsx')
            except Exception:
                self.pattern_wb.save('files/' + str(i) + '.xlsx')
            
            if i % 5 == 0:
                self.progress_bar.setValue(int((i+1) / self.source_ws.max_row * 100))
                
        self.progress_bar.setValue(100)
        self.refresh_table()
        
        msgWarning = QMessageBox()
        msgWarning.setText('Все файлы были созданы'); 
        msgWarning.setIcon(QMessageBox.Icon.NoIcon);
        msgWarning.setWindowTitle("Успех");
        msgWarning.setWindowIcon(self.icon) 
        msgWarning.exec();
        self.progress_bar.setValue(0)

                    
    def refresh_table(self):
        for sheet_name, sheet_model in self.pattern_sheets.items():
            try:
                _ = self.excels_default_value[sheet_name]
            except KeyError:
                self.excels_default_value[sheet_name] = {}
                
            default_dict = self.excels_default_value[sheet_name]
            for key in default_dict.keys():
                beg_cell = coordinate_from_string(key)
                
                cell_row = beg_cell[1] - 1
                cell_column = column_index_from_string(beg_cell[0]) - 1
                cell_index = sheet_model.index(cell_row, cell_column)
                
                cell_value = QVariant(None) 
                sheet_model.setData(cell_index, cell_value, Qt.DisplayRole)
                sheet_model.addIndecesDefaultValue(cell_index)
            
            try:
                _ = self.excels_relation[sheet_name]
            except KeyError:
                self.excels_relation[sheet_name] = {}
                
            sheet_dict = self.excels_relation[sheet_name]
            for key, value in sheet_dict.items():
                beg_cell = coordinate_from_string(key)
                
                cell_row = beg_cell[1] - 1
                cell_column = column_index_from_string(beg_cell[0]) - 1
                cell_index = sheet_model.index(cell_row, cell_column)
                
                cell_value = ''
                if len(self.source_col_names) < column_index_from_string(value) + 1:
                    cell_value = QVariant(value)
                else:
                    cell_value = QVariant(self.source_col_names[column_index_from_string(value)]) 
                sheet_model.setData(cell_index, cell_value, Qt.DisplayRole)
                sheet_model.addIndecesConnExcel(cell_index)
                
            try:
                _ = self.excels_relation_many[sheet_name]
            except KeyError:
                self.excels_relation_many[sheet_name] = {}

            sheet_many_dict = self.excels_relation_many[sheet_name]
            for key, value in sheet_many_dict.items():
                beg_cell = coordinate_from_string(key)
                
                cell_row = beg_cell[1] - 1
                cell_column = column_index_from_string(beg_cell[0]) - 1
                cell_index = sheet_model.index(cell_row, cell_column)
                
                items = value['delimeter'].join(x for x in value['items'])
                cell_value = QVariant(items) 
                sheet_model.setData(cell_index, cell_value, Qt.DisplayRole)
                sheet_model.addIndecesConnExcel(cell_index)

            sheet_model.endResetModel()

    def action_about_handler(self):
        self.license.show()
        
    def action_new_handler(self):
        self.reset()
        self.refresh_list_items()
        self.refresh_tabs()
        self.progress_bar.setValue(0)
        self.refresh_table()
                
    def action_save_handler(self):
        if self.saves_path != '' and QFileInfo(self.saves_path).exists():
            self.save(self.saves_path)
        else:
            self.action_save_as_handler() 
               
    def action_save_as_handler(self):
        Path('saves').mkdir(parents=True, exist_ok=True)
        path, _ = QFileDialog.getSaveFileName(self, 'Введите имя сохранения', 'saves')
        if path != '':
            self.save(path)
                
    def save(self, save_path: str):
        if not save_path.endswith(".json"):
            save_path = save_path + '.json'
        self.saves_path = save_path
        config = ConfigParser()
        config.add_section('Info')
        config.set('Info', 'recent_save', save_path)
        with open('settings.ini', 'w', encoding='utf-8') as configfile:
            config.write(configfile)
        
        data = {
            'source_file' : self.source_path,
            'pattern_file' : self.pattern_path,
            'default_values' : self.excels_default_value,
            'excels_connection' : self.excels_relation,
            'excels_relation_many' : self.excels_relation_many,
            'col_index_name' : str(self.combo_box_cell.currentIndex()),
            'add_name' : str(self.line_edit_added.text()),
            'check_name' : str(self.check_box_added.isChecked())
            }
            
        with open(save_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
                
    def action_load_handler(self, file=None):
        if file is None or isinstance(file, bool):
            path, _ = QFileDialog.getOpenFileName(self, 'Откройте сохранение', 'saves', '*.json')
            if path == '':
                return
        else:
            path = str(file)
             
        with open(path, encoding='utf-8') as f:
            save = json.load(f)

            try:
                _ = save['source_file']
                _ = save['pattern_file']
                _ = save['default_values']
                _ = save['excels_connection']
                _ = save['excels_relation_many']
                _ = save['col_index_name']
                _ = save['add_name']
                _ = save['check_name']
            except KeyError:
                self.showWarning("Неподходящее сохранение")
                return
            
            self.reset()
            
            sourceLoaded = True
            if(not self.process_source_file(Path(save['source_file']).name, False)):
                if not self.process_source_file(save['source_file']):
                    sourceLoaded = False
                
            patternLoaded = True
            if(not self.process_pattern_file(Path(save['pattern_file']).name, False)):
                if not self.process_pattern_file(save['pattern_file']):
                    patternLoaded = False
                
            # self.process_pattern_file(save['pattern_file'])
            if patternLoaded:
                self.excels_default_value = save['default_values']
                self.excels_relation = save['excels_connection']
                self.excels_relation_many = save['excels_relation_many']
            
            self.refresh_list_items()
            self.refresh_tabs()
            self.progress_bar.setValue(0)
            self.refresh_table()
            
            if sourceLoaded:
                index = int(save['col_index_name'])
                if index >= 0:
                    self.combo_box_cell.setCurrentIndex(index)
                
            self.line_edit_added.setText(save['add_name'])
            self.check_box_added.setChecked(save['check_name'] == "True")
            
            self.saves_path = str(Path(path).resolve())
            self.save(self.saves_path)
                
    def reset(self):
        self.saves_path = ''
        self.source_path = ''
        self.pattern_path = ''
        self.excels_default_value = {}
        self.excels_relation = {}
        self.excels_relation_many = {}
        
        self.pattern_sheets = {}
        self.source_col_names = []
        self.current_source_columns = ''
        self.orig_pattern_wb = ''
        self.pattern_wb = ''
        self.active_sheet_name = ''
        self.active_cell = []
        self.tabs_names = []
        self.source_ws=''
         
        self.progress_bar.setValue(0)
        self.line_edit_source.setText('')
        self.line_edit_pattern.setText('')
        
        self.table_view.setModel(SheetModel(Workbook().create_sheet("Пустой")))
        
    def showWarning(self, message):
        msgWarning = QMessageBox()
        msgWarning.setText("ПРЕДУПРЕЖДЕНИЕ!\n" + message);
        msgWarning.setIcon(QMessageBox.Icon.Warning);
        msgWarning.setWindowTitle("Внимание");
        msgWarning.setWindowIcon(self.icon) 
        msgWarning.exec();

def main():    
    app = QApplication(sys.argv)
    wnd = Window()
    wnd.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
    
